import xlrd as dadosExcel;
from openpyxl import load_workbook
from datetime import datetime
import re as re;

class RecebeDados:
        
    def __init__(self, plan_dados):
        # abre a planilha a ser usada para receber os dados
        # self.planilha = dadosExcel.open_workbook(plan_dados);
        print(f"Carregando a planilha {plan_dados}")
        start = datetime.now()
        self.planilha = load_workbook(filename = plan_dados, data_only=True)
        print(f"Planilha carregada em {datetime.now() - start}")
        return;
    
    def defineAba(self, nomeAba):
        # abre a aba da planilha a ser usada para receber os dados
        # self.aba = self.planilha.sheet_by_name(nomeAba);
        self.aba = self.planilha[nomeAba]
        return;
    
    def pegaEscalar (self, celula, lin_offset=0, col_offset=0):
        # converte do formato A1 para 1,1
        coord = self.cell2num(celula);
        
        # verifica se a linha a ser acessada existe na planilha
        # if ((coord[0] +lin_offset)>self.aba.nrows): 
        if ((coord[0] +lin_offset)>self.aba.max_row): 
            return None;
        
        # verifica se a coluna a ser acessada existe na planilha
        # elif ((coord[1]+col_offset)>self.aba.ncols):
        elif ((coord[1]+col_offset)>self.aba.max_column):
            return None;
            
        else:
            # acessa o valor da planilha
            # resultado = self.aba.cell_value(coord[0]-1+lin_offset, coord[1]-1+col_offset);
            resultado = self.aba.cell(row=coord[0]+lin_offset, column=coord[1]+col_offset).value
            
            # se for uma string vazia retorna None
            if (resultado == ""):
                return None;
            else:
                return resultado;
    
    def pegaVetor (self, celula, direcao, tamanho, lin_offset=0, col_offset=0):
        
        coord =self.cell2num(celula);
        
        # if ((coord[0] +lin_offset)>self.aba.nrows): 
        if ((coord[0] +lin_offset)>self.aba.max_row): 
            return None;
        # elif ((coord[1]+col_offset)>self.aba.ncols):
        elif ((coord[1]+col_offset)>self.aba.max_column):
            return None;
        else:
            if (direcao == 'horizontal'):
                # retorna o valor da celula mantendo a linha fixa e variando a coluna
                # return self.aba.row_values(int(coord[0]-1+lin_offset), int(coord[1]-1+col_offset), int(coord[1]-1+col_offset+tamanho));
                return [self.aba.cell(row=int(coord[0]+lin_offset), column=i).value for i in range(int(coord[1]+col_offset), int(coord[1]+col_offset+tamanho))]
            elif (direcao == 'vertical'):
                # retorna o valor da celula mantendo a coluna fixa e variando a linha
                # return self.aba.col_values(int(coord[1]-1+col_offset), int(coord[0]-1+lin_offset), int(coord[0]-1+lin_offset+tamanho));
                return [self.aba.cell(row=i, column=int(coord[1]+col_offset)).value for i in range(int(coord[0]+lin_offset), int(coord[0]+lin_offset+tamanho))]
            else:
                print("direcao %s nao permitida" % direcao);
                return;
    
    def cell2num(self, celula):
        # funcao responsavel por converter do formato de A1 para 1,1        
        # pega a parte literal na string celula (coluna)
        match = re.search("[A-Z]+", celula.upper());
        coluna_str = match.group();
        
        # pega a parte numerica (linha)
        match = re.search("[0-9]+", celula);
        linha = int(match.group());
        
        # converte a parte literal em numero
        # inicialmente pega o numero de caracteres
        max_exp = len(coluna_str);
        coluna=0;
        
        # percorre todos os caracteres
        for i_exp in range(max_exp, 0, -1):
            # computa o numero da coluna (converte para ascii e potencializa para retornar na base 10)
            coluna = int(coluna + (26**(i_exp-1))*(ord(coluna_str[max_exp-i_exp])-64));
        
        # salva na lista a linha e coluna correntes (nao eh uma boa solucao, mas foi feito para nao passar parametro por referencia)
        lista = [linha, coluna];
       
        return lista;
        
